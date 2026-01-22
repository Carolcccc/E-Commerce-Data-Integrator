#!/usr/bin/env python3
"""
Shopee to Myship Product Import
Generates proper format for single or dual variant products
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import shutil
import os
import json

print("=" * 80)
print("SHOPEE TO MYSHIP PRODUCT IMPORT CONVERSION")
print("=" * 80)

# ============================================================
# LOAD CONFIGURATION
# ============================================================
print("\nLoading configuration...")
try:
    with open('config.json', 'r', encoding='utf-8') as f:
        config = json.load(f)
    print("Configuration loaded from config.json")
except FileNotFoundError:
    print("ERROR: config.json not found. Please copy config.example.json to config.json and fill in your details.")
    exit(1)

# ============================================================
# 1. LOAD SOURCE DATA
# ============================================================
print("\nLoading source files...")

sales_file = config['files']['sales']
basicinfo_file = config['files']['basicinfo']
media_file = config['files']['media']

sales = pd.read_excel(sales_file, dtype={'et_title_product_id': str}, engine='calamine', header=0)
sales = sales[sales['et_title_product_id'].notna() & 
              (sales['et_title_product_id'] != 'sales_info') &
              (sales['et_title_product_id'] != 'productID')]

basicinfo = pd.read_excel(basicinfo_file, dtype={'et_title_product_id': str}, engine='calamine', header=0)
basicinfo = basicinfo[basicinfo['et_title_product_id'].notna() & 
                      (basicinfo['et_title_product_id'] != 'basic_info') &
                      (basicinfo['et_title_product_id'] != 'productID')]

media = pd.read_excel(media_file, dtype={'et_title_product_id': str}, engine='calamine', header=0)

sales_file = config['files']['sales']
basicinfo_file = config['files']['basicinfo']
media_file = config['files']['media']

sales = pd.read_excel(sales_file, dtype={'et_title_product_id': str}, engine='calamine', header=0)
sales = sales[sales['et_title_product_id'].notna() & 
              (sales['et_title_product_id'] != 'sales_info') &
              (sales['et_title_product_id'] != 'productID')]

basicinfo = pd.read_excel(basicinfo_file, dtype={'et_title_product_id': str}, engine='calamine', header=0)
basicinfo = basicinfo[basicinfo['et_title_product_id'].notna() & 
                      (basicinfo['et_title_product_id'] != 'basic_info') &
                      (basicinfo['et_title_product_id'] != 'productID')]

media = pd.read_excel(media_file, dtype={'et_title_product_id': str}, engine='calamine', header=0)
media = media[media['et_title_product_id'].notna() & 
              (media['et_title_product_id'] != 'media_info') &
              (media['et_title_product_id'] != 'productID')]

print(f"Loaded sales: {len(sales)} rows")
print(f"Loaded basicinfo: {len(basicinfo)} rows")
print(f"Loaded media: {len(media)} rows")

# ============================================================
# 2. MERGE DATA
# ============================================================
print("\nMerging data...")
merged = sales.merge(basicinfo, on='et_title_product_id', how='left', suffixes=('', '_basic'))
merged = merged.merge(media, on='et_title_product_id', how='left', suffixes=('', '_media'))
print(f"Merged: {len(merged)} rows")

# ============================================================
# 3. TRANSFORM TO MYSHIP FORMAT
# ============================================================
print("\nTransforming to Myship product format...")

# Create output dataframe with required columns for single product input
output = pd.DataFrame()

# Required columns according to template
output['Product Name'] = merged['et_title_product_name'].fillna('')

# Convert image hash to URL
def convert_to_url(hash_id):
    if pd.isna(hash_id) or hash_id == '':
        return ''
    hash_str = str(hash_id).replace('https://s-cf-tw.shopeesz.com/file/', '')
    return f'https://s-cf-tw.shopeesz.com/file/{hash_str}'

if 'ps_item_cover_image' in merged.columns:
    output['Product Image (URL)'] = merged['ps_item_cover_image'].apply(convert_to_url).fillna('')
else:
    output['Product Image (URL)'] = ''

output['Product Description'] = merged['et_title_product_description'].fillna('')

# Variant name
output['Variant'] = merged['et_title_variation_name'].fillna('Default')

# Quantity from stock
output['Quantity'] = pd.to_numeric(merged['et_title_variation_stock'], errors='coerce').fillna(0).astype(int)

# Price
output['Price'] = pd.to_numeric(merged['et_title_variation_price'], errors='coerce').fillna(0).astype(int)

# Optional fields
output['Discount Price'] = ''
output['Product Status'] = 'New'
output['Order Limit'] = ''
output['Minimum Order'] = ''
output['Verification Status'] = ''
output['Upload Status'] = ''

print(f"Transformed {len(output)} products")

# ============================================================
# 4. CREATE BATCH FILES (ALL PRODUCTS, 100 PER FILE)
# ============================================================
batch_size = config['batch']['size']
max_file_size_mb = config['batch']['max_file_size_mb']
total_products = len(output)
num_batches = (total_products + batch_size - 1) // batch_size

print(f"\nCreating {num_batches} batch file(s) of up to {batch_size} products each...")

template_file = config['files']['template']
start_row = 7

for batch_index in range(num_batches):
    batch_start = batch_index * batch_size
    batch_end = min(batch_start + batch_size, total_products)
    batch_df = output.iloc[batch_start:batch_end].copy().reset_index(drop=True)

    output_file = f'myship_upload_batch{batch_index + 1}.xlsm'
    shutil.copy(template_file, output_file)

    wb = load_workbook(output_file, keep_vba=True)
    sheet = wb['Single Product Import']

    # Clear existing data (rows 7+)
    for row_idx in range(start_row, sheet.max_row + 1):
        for col_idx in range(1, 14):
            sheet.cell(row=row_idx, column=col_idx).value = None

    # Write data starting from row 7
    for row_offset, row_data in batch_df.iterrows():
        row_num = start_row + row_offset
        for col_idx, col_name in enumerate(batch_df.columns, start=1):
            value = row_data[col_name]
            if pd.isna(value):
                value = ''
            sheet.cell(row=row_num, column=col_idx).value = value

    wb.save(output_file)

    file_size_mb = os.path.getsize(output_file) / (1024 * 1024)
    size_check = "OK" if file_size_mb < max_file_size_mb else "EXCEEDS LIMIT"

    print(f"\nBatch {batch_index + 1}/{num_batches} created")
    print(f"  File: {output_file}")
    print(f"  Products: {len(batch_df)}")
    print(f"  Size: {file_size_mb:.2f} MB ({size_check})")
    print(f"  Sheets: {wb.sheetnames}")

    print(f"  Summary:")
    print(f"    Unique product names: {batch_df['Product Name'].nunique()}")
    print(f"    Total quantity: {batch_df['Quantity'].sum()}")
    print(f"    Price range: {batch_df['Price'].min()} - {batch_df['Price'].max()}")
    print(f"    Products with images: {(batch_df['Product Image (URL)'] != '').sum()}")
