#!/usr/bin/env python3
"""
Detailed check for product name formatting issues
Validates special characters and length constraints
"""

import openpyxl
import unicodedata

print("=" * 80)
print("DETAILED PRODUCT NAME VALIDATION")
print("=" * 80)

try:
    wb = openpyxl.load_workbook('myship_upload_batch1.xlsm', keep_vba=True)
except FileNotFoundError:
    print("ERROR: myship_upload_batch1.xlsm not found")
    exit(1)

sheet = wb['Single Product Import']

# Special characters that need conversion
special_chars = {
    '&': 'Full-width &',
    '(': 'Full-width (',
    ')': 'Full-width )',
    '=': 'Full-width =',
    ';': 'Full-width ;',
    "'": 'Full-width apostrophe',
    '"': 'Full-width quote',
    '<': 'Full-width <',
    '>': 'Full-width >',
    '\\': 'Full-width backslash'
}

print("\nChecking first 10 product names:")
print("-" * 80)

issues_found = []

for row_idx in range(7, 17):
    name = sheet.cell(row=row_idx, column=1).value
    if not name:
        break
    
    product_num = row_idx - 6
    print(f"\nProduct {product_num}:")
    print(f"  Name: {name}")
    print(f"  Length: {len(name)} characters")
    
    has_issue = False
    for i, char in enumerate(name):
        # Check for half-width special characters
        if char in special_chars:
            has_issue = True
            issues_found.append((product_num, row_idx, char, special_chars[char]))
            print(f"  ERROR at position {i}: '{char}' should be full-width")
        
        # Check for control characters
        cat = unicodedata.category(char)
        if cat in ['Cc', 'Cf', 'Cs', 'Co', 'Cn']:
            has_issue = True
            print(f"  WARNING at position {i}: Control character U+{ord(char):04X}")
    
    # Check length
    if len(name) > 60:
        has_issue = True
        issues_found.append((product_num, row_idx, 'LENGTH', len(name)))
        print(f"  ERROR: Exceeds 60 character limit ({len(name)} characters)")
    
    if not has_issue:
        print(f"  OK: No issues found")

print("\n" + "=" * 80)
print("VALIDATION SUMMARY")
print("=" * 80)

if issues_found:
    print(f"\nFound {len(issues_found)} issue(s) to fix:\n")
    
    # Categorize issues
    char_issues = [x for x in issues_found if x[2] != 'LENGTH']
    length_issues = [x for x in issues_found if x[2] == 'LENGTH']
    
    if char_issues:
        print(f"Character Issues ({len(char_issues)} total):")
        for product_num, row_idx, char, description in char_issues:
            print(f"  Product {product_num} (Row {row_idx}): '{char}' needs to be {description}")
    
    if length_issues:
        print(f"\nLength Issues ({len(length_issues)} total):")
        for product_num, row_idx, _, length in length_issues:
            print(f"  Product {product_num} (Row {row_idx}): {length} characters (exceeds 60 character limit)")
    
    print(f"\nRecommendation: Review and fix the issues above")
else:
    print("\nOK: All product names meet format requirements!")

print("\n" + "=" * 80)
print("CHARACTER ENCODING CHECK")
print("=" * 80)

sample_name = sheet.cell(row=7, column=1).value
if sample_name:
    print(f"\nSample product name: {sample_name}")
    print(f"\nUTF-8 Encoding Test:")
    try:
        encoded = sample_name.encode('utf-8')
        print(f"  OK: Successfully encoded to UTF-8 ({len(encoded)} bytes)")
    except UnicodeEncodeError as e:
        print(f"  ERROR: Cannot encode to UTF-8: {e}")

print("\n" + "=" * 80)
