#!/usr/bin/env python3
"""
Check for special character format compliance
Validates that all special characters are in full-width format
"""

import openpyxl

print("=" * 80)
print("SPECIAL CHARACTER FORMAT VALIDATION")
print("=" * 80)

# Half-width to full-width character mapping
special_chars = {
    '&': 'Full-width &',
    '(': 'Full-width (',
    ')': 'Full-width )',
    '=': 'Full-width =',
    ';': 'Full-width ;',
    "'": 'Full-width apostrophe',
    '"': 'Full-width quote',
    '<': 'Full-width <',
    '>': 'Full-width >',
    '\\': 'Full-width backslash'
}

try:
    wb = openpyxl.load_workbook('myship_upload_test10.xlsm', keep_vba=True)
except FileNotFoundError:
    print("ERROR: myship_upload_test10.xlsm not found")
    exit(1)

sheet = wb['Single Product Import']

print("\nValidating fields: Product Name, Product Description, Variant")
print("-" * 80)

issues = []

# Check first 10 products
for row_idx in range(7, 17):
    name = sheet.cell(row_idx, 1).value
    desc = sheet.cell(row_idx, 3).value
    variant = sheet.cell(row_idx, 4).value
    
    if not name:
        break
    
    product_num = row_idx - 6
    
    # Check product name
    for half_char in special_chars.keys():
        if half_char in str(name):
            issues.append(f"Product {product_num} - Product Name contains half-width '{half_char}': {name[:50]}")
    
    # Check product description
    for half_char in special_chars.keys():
        if half_char in str(desc if desc else ''):
            issues.append(f"Product {product_num} - Product Description contains half-width '{half_char}'")
    
    # Check variant
    for half_char in special_chars.keys():
        if half_char in str(variant if variant else ''):
            issues.append(f"Product {product_num} - Variant contains half-width '{half_char}': {variant}")

if issues:
    print(f"\nERROR: Found {len(issues)} special character issue(s):\n")
    for issue in issues:
        print(f"  {issue}")
else:
    print("\nOK: All fields contain only proper full-width special characters")

# Display examples
print("\n" + "-" * 80)
print("First 3 Products - Sample:")
print("-" * 80)

for row_idx in range(7, 10):
    name = sheet.cell(row_idx, 1).value
    desc = sheet.cell(row_idx, 3).value
    
    print(f"\nProduct {row_idx - 6}:")
    print(f"  Name: {name}")
    print(f"  Description (first 100 chars): {str(desc if desc else '')[:100]}...")
    
    # Check for any special characters
    has_special = False
    for half_char in special_chars.keys():
        if half_char in str(name) or half_char in str(desc if desc else ''):
            has_special = True
            print(f"  WARNING: Contains problematic character: {half_char}")
    
    if not has_special:
        print(f"  OK: No formatting issues detected")

print("\n" + "=" * 80)
