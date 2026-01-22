#!/usr/bin/env python3
"""
Create proper Myship upload file with store information
Fills in store sheet first, then products reference that store
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import json

print("=" * 80)
print("CREATING PROPER MYSHIP UPLOAD FILE")
print("=" * 80)

# ============================================================
# LOAD CONFIGURATION
# ============================================================
print("\nLoading configuration...")
try:
    with open('config.json', 'r', encoding='utf-8') as f:
        config = json.load(f)
    print("Configuration loaded from config.json")
except FileNotFoundError:
    print("ERROR: config.json not found. Please copy config.example.json to config.json and fill in your details.")
    exit(1)

# ============================================================
# 1. LOAD SOURCE DATA
# ============================================================
print("\nLoading source files...")

sales = pd.read_excel(config['files']['sales'], dtype={'et_title_product_id': str}, engine='calamine', header=0)
sales = sales[sales['et_title_product_id'].notna() & 
              (sales['et_title_product_id'] != 'sales_info') &
              (sales['et_title_product_id'] != 'product_id')]

basicinfo = pd.read_excel(config['files']['basicinfo'], dtype={'et_title_product_id': str}, engine='calamine', header=0)
basicinfo = basicinfo[basicinfo['et_title_product_id'].notna() & 
                      (basicinfo['et_title_product_id'] != 'basic_info') &
                      (basicinfo['et_title_product_id'] != 'product_id')]

media = pd.read_excel(config['files']['media'], dtype={'et_title_product_id': str}, engine='calamine', header=0)
media = media[media['et_title_product_id'].notna() & 
              (media['et_title_product_id'] != 'media_info') &
              (media['et_title_product_id'] != 'product_id')]

print(f"Loaded sales: {len(sales)} rows")
print(f"Loaded basicinfo: {len(basicinfo)} rows") 
print(f"Loaded media: {len(media)} rows")

# ============================================================
# 2. MERGE DATA
# ============================================================
merged = sales.merge(basicinfo, on='et_title_product_id', how='left', suffixes=('', '_basic'))
merged = merged.merge(media, on='et_title_product_id', how='left', suffixes=('', '_media'))

# ============================================================
# 3. PREPARE STORE INFO
# ============================================================
store_name = config['store']['name']
store_description = config['store']['description']
store_temperature = config['store']['temperature']

# Get first image for store
if 'ps_item_cover_image' in media.columns and len(media) > 0:
    first_image_hash = media.iloc[0]['ps_item_cover_image']
    if pd.notna(first_image_hash):
        store_image = f'https://s-cf-tw.shopeesz.com/file/{first_image_hash}'
    else:
        store_image = ''
else:
    store_image = ''

print(f"\nStore info:")
print(f"  Name: {store_name}")
print(f"  Description: {store_description[:50]}...")

# ============================================================
# 4. LOAD TEMPLATE AND FILL STORE SHEET
# ============================================================
template_file = config['files']['template']
output_file = 'myship_upload_ready.xlsm'

print(f"\nLoading template: {template_file}")
wb = load_workbook(template_file, keep_vba=True)

print("\nFilling store sheet...")
store_sheet = wb['Store Import']

# Clear existing data (rows 7+)
for row_idx in range(7, store_sheet.max_row + 1):
    for col_idx in range(1, store_sheet.max_column + 1):
        store_sheet.cell(row=row_idx, column=col_idx).value = None

# Fill store info in row 7
store_sheet.cell(row=7, column=1).value = store_temperature
store_sheet.cell(row=7, column=2).value = store_name
store_sheet.cell(row=7, column=3).value = store_description
store_sheet.cell(row=7, column=4).value = store_image
store_sheet.cell(row=7, column=5).value = ''
store_sheet.cell(row=7, column=6).value = ''

print(f"Store sheet filled with: {store_name}")

# ============================================================
# 5. FILL PRODUCT SHEET
# ============================================================
print("\nFilling product sheet...")
product_sheet = wb['Single Product Import']

# Clear existing data (rows 7+)
for row_idx in range(7, product_sheet.max_row + 1):
    for col_idx in range(1, 14):
        product_sheet.cell(row=row_idx, column=col_idx).value = None

# Transform data
output = pd.DataFrame()
output['Product Name'] = merged['et_title_product_name'].fillna('')

def convert_to_url(hash_id):
    if pd.isna(hash_id) or hash_id == '':
        return ''
    img_hash_str = str(hash_id).replace('https://s-cf-tw.shopeesz.com/file/', '')
    return f'https://s-cf-tw.shopeesz.com/file/{img_hash_str}'

if 'ps_item_cover_image' in merged.columns:
    output['Product Image (URL)'] = merged['ps_item_cover_image'].apply(convert_to_url).fillna('')
else:
    output['Product Image (URL)'] = ''

output['Product Description'] = merged['et_title_product_description'].fillna('')
output['Variant'] = merged['et_title_variation_name'].fillna('Default')
output['Quantity'] = pd.to_numeric(merged['et_title_variation_stock'], errors='coerce').fillna(0).astype(int)
output['Price'] = pd.to_numeric(merged['et_title_variation_price'], errors='coerce').fillna(0).astype(int)
output['Discount Price'] = ''
output['Product Status'] = 'New'
output['Order Limit'] = ''
output['Minimum Order'] = ''
output['Store Name'] = store_name
output['Verification Status'] = ''
output['Upload Status'] = ''

# Write to sheet
for row_offset, row_data in output.iterrows():
    row_num = 7 + row_offset
    for col_idx, col_name in enumerate(output.columns, start=1):
        value = row_data[col_name]
        if pd.isna(value):
            value = ''
        product_sheet.cell(row=row_num, column=col_idx).value = value

print(f"Added {len(output)} products")

# ============================================================
# 6. SAVE FILE
# ============================================================
wb.save(output_file)
file_size_mb = os.path.getsize(output_file) / (1024 * 1024)

print(f"\nFile created successfully!")
print(f"  Output: {output_file}")
print(f"  Size: {file_size_mb:.2f} MB")
print(f"  Sheets: {wb.sheetnames}")
print(f"\nSummary:")
print(f"  Store: {store_name}")
print(f"  Products: {len(output)}")
print(f"  Total quantity: {output['Quantity'].sum()}")
print(f"  Price range: {output['Price'].min()} - {output['Price'].max()}")
